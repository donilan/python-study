#!/usr/bin/env python
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter

if __name__ == '__main__':
  wb = Workbook()
  dest_filename = r'empty_book.xlsx'
  ws = wb.worksheets[0]
  ws.title = "range names"
  for col_idx in xrange(1, 40):
    col = get_column_letter(col_idx)
    for row in xrange(1, 600):
      ws.cell('%s%s'%(col, row)).value = '%s%s' % (col, row)

  ws = wb.create_sheet()

  ws.title = 'Pi'

  ws.cell('F5').value = 3.14

  wb.save(filename = dest_filename)
